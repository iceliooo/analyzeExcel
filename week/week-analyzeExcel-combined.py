import pandas as pd
from datetime import datetime, timedelta
import argparse
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def parse_args():
    """解析命令行参数，支持时间区间输入"""
    # 获取当前日期和30天前的日期作为默认值
    default_end_date = datetime.now().strftime('%Y-%m-%d')
    default_start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    # 从控制台获取开始日期和结束日期
    start_date = input(f'请输入开始日期 (YYYY-MM-DD) [默认: {default_start_date}]: ')
    end_date = input(f'请输入结束日期 (YYYY-MM-DD) [默认: {default_end_date}]: ')
    
    # 如果用户未输入,使用默认值
    if not start_date:
        start_date = default_start_date
    if not end_date:
        end_date = default_end_date

    parser = argparse.ArgumentParser(description='分析Excel数据，支持时间区间过滤')
    parser.add_argument('--start_date', type=str, default=start_date,
                      help='开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end_date', type=str, default=end_date,
                      help='结束日期 (YYYY-MM-DD)') 
    parser.add_argument('--excel_file', type=str, default='file/0711/0711V8客户问题上报 - 分析.xlsx',
                      help='Excel文件路径')
    return parser.parse_args()

def format_excel_worksheet(worksheet, summary_df):
    """格式化Excel工作表"""
    # 定义填充颜色
    blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    orange_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    
    # 定义换行对齐方式
    wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    # 需要换行显示的列
    wrap_columns = ['有效问题总处理时长(小时)', '有效问题平均处理时长(小时)']
    wrap_col_indexes = []
    
    # 获取需要换行的列索引
    for i, col_name in enumerate(summary_df.columns):
        if col_name in wrap_columns:
            wrap_col_indexes.append(i + 2)  # +2 因为Excel索引从1开始且有索引列
    
    # 为所有列标题设置蓝色背景和居中对齐
    for col in range(1, len(summary_df.columns) + 2):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置适应内容的列宽
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].auto_size = True
    
    # 为需要换行的列设置格式
    for col_idx in wrap_col_indexes:
        # 为标题设置换行对齐
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.alignment = wrap_alignment
        
        # 为所有数据行设置换行对齐
        for row in range(2, len(summary_df) + 2):  # +2 因为Excel有标题行和索引从1开始
            cell = worksheet.cell(row=row, column=col_idx)
            cell.alignment = wrap_alignment
    
    # 查找特定列的索引
    special_columns = ['有效问题平均处理时长(小时)', '处理及时率%', '一次交付率%', '严重问题率%']
    special_col_indexes = []
    
    # 获取列索引 (注意Excel中的列从1开始，且第1列是索引)
    for i, col_name in enumerate(summary_df.columns):
        if col_name in special_columns:
            special_col_indexes.append(i + 2)  # +2 因为Excel索引从1开始且有索引列
    
    # 为特定列设置橙色背景
    for col_idx in special_col_indexes:
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.fill = orange_fill
        
        # 为总计行的特定单元格设置橙色
        total_row_cell = worksheet.cell(row=len(summary_df) + 1, column=col_idx)
        total_row_cell.fill = orange_fill
    
    # 自动调整列宽以适应内容
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def analyze_by_customer(df, report_time_mask, solve_time_mask):
    """按客户名称分组统计"""
    print("\n正在分析客户维度数据...")
    
    # 按客户名称分组统计
    # 1. 统计每个客户的总问题量（按上报时间）
    total_by_customer = df[report_time_mask]['KH_客户名称'].value_counts()

    # 2. 统计每个客户的处理完成量（按解决时间）
    completed_by_customer = df[
        (solve_time_mask) & (df['开发处理-诊断结论-二级'].notna())
    ]['KH_客户名称'].value_counts()

    # 3. 统计每个客户的有效问题量（按解决时间）
    valid_by_customer = df[
        (solve_time_mask) & (df['有效问题'] == '是')
    ]['KH_客户名称'].value_counts()

    # 4. 统计每个客户的缺陷问题量（按解决时间）
    defect_by_customer = df[
        (solve_time_mask) & (df['缺陷问题'] == '是') 
    ]['KH_客户名称'].value_counts()

    # 5. 统计每个客户的开发超期量（按解决时间）
    overdue_by_customer = df[
        (solve_time_mask) & (df['开发是否超期'] == '是')
    ]['KH_客户名称'].value_counts()

    # 6. 统计每个客户的有效问题处理时长（按解决时间）
    valid_issues = df[(solve_time_mask) & (df['有效问题'] == '是')]
    total_time_by_customer = valid_issues.groupby('KH_客户名称')['开发处理-投入工作量'].sum()
    # 计算平均处理时长
    avg_time_by_customer = valid_issues.groupby('KH_客户名称')['开发处理-投入工作量'].mean()

    # 7. 统计每个客户的一次交付量（按解决时间）
    one_delivery_by_customer = df[
        (solve_time_mask) & (df['一次交付'] == '是')
    ]['KH_客户名称'].value_counts()

    # 8. 统计每个客户的严重问题量（按解决时间）
    serious_by_customer = df[
        (solve_time_mask) & (df['严重问题'] == '是')
    ]['KH_客户名称'].value_counts()

    # 将结果存储在字典中
    result = {
        'total_count': total_by_customer,
        'completed_count': completed_by_customer,
        'valid_count': valid_by_customer,
        'defect_count': defect_by_customer,
        'overdue_count': overdue_by_customer,
        'total_time': total_time_by_customer,
        'avg_time': avg_time_by_customer,
        'one_delivery_count': one_delivery_by_customer,
        'serious_count': serious_by_customer
    }

    return create_summary_dataframe(result, "客户")

def analyze_by_developer(df, report_time_mask, solve_time_mask):
    """按开发人员分组统计"""
    print("\n正在分析开发人员维度数据...")
    
    # 将开发处理_开发人员的空值替换为"未知"
    df['开发处理_开发人员'] = df['开发处理_开发人员'].fillna('未知')

    # 按开发人员分组统计
    # 1. 统计每个开发人员的总问题量（按上报时间）
    total_by_developer = df[report_time_mask]['开发处理_开发人员'].value_counts()

    # 2. 统计每个开发人员的有效问题量（按解决时间）
    valid_by_developer = df[
        (solve_time_mask) & (df['有效问题'] == '是')
    ]['开发处理_开发人员'].value_counts()

    # 3. 统计每个开发人员的缺陷问题量（按解决时间）
    defect_by_developer = df[
        (solve_time_mask) & (df['缺陷问题'] == '是') 
    ]['开发处理_开发人员'].value_counts()

    # 4. 统计每个开发人员的开发超期量（按解决时间）
    overdue_by_developer = df[
        (solve_time_mask) & (df['开发是否超期'] == '是')
    ].apply(lambda x: x['开发处理-超期责任人'] if pd.notna(x['开发处理-超期责任人']) else x['开发处理_开发人员'], axis=1).value_counts()

    # 5. 统计每个开发人员的有效问题处理时长（按解决时间）
    valid_issues = df[(solve_time_mask) & (df['有效问题'] == '是')]
    total_time_by_developer = valid_issues.groupby('开发处理_开发人员')['开发处理-投入工作量'].sum()
    # 计算平均处理时长
    avg_time_by_developer = valid_issues.groupby('开发处理_开发人员')['开发处理-投入工作量'].mean()

    # 6. 统计每个开发人员的一次交付量（按解决时间）
    one_delivery_by_developer = df[
        (solve_time_mask) & (df['一次交付'] == '是')
    ]['开发处理_开发人员'].value_counts()

    # 7. 统计每个开发人员的严重问题量（按解决时间）
    serious_by_developer = df[
        (solve_time_mask) & (df['严重问题'] == '是')
    ]['开发处理_开发人员'].value_counts()

    # 8. 统计每个开发人员的处理完成量（按解决时间）
    completed_by_developer = df[
        (solve_time_mask) & (df['开发处理-诊断结论-二级'].notna())
    ]['开发处理_开发人员'].value_counts()

    # 将结果存储在字典中
    result = {
        'total_count': total_by_developer,
        'completed_count': completed_by_developer,
        'valid_count': valid_by_developer,
        'defect_count': defect_by_developer,
        'overdue_count': overdue_by_developer,
        'total_time': total_time_by_developer,
        'avg_time': avg_time_by_developer,
        'one_delivery_count': one_delivery_by_developer,
        'serious_count': serious_by_developer
    }

    return create_summary_dataframe(result, "开发人员")

def analyze_by_department(df, report_time_mask, solve_time_mask):
    """按研发部门分组统计"""
    print("\n正在分析部门维度数据...")
    
    # 将开发处理-开发所属部门的空值替换为"未知"
    df['开发处理-开发所属部门'] = df['开发处理-开发所属部门'].fillna('未知')

    # 按研发组分组统计
    # 1. 统计每个组的总问题量（按上报时间）
    total_by_group = df[report_time_mask]['开发处理-开发所属部门'].value_counts()

    # 2. 统计每个组的有效问题量（按解决时间）
    valid_by_group = df[
        (solve_time_mask) & (df['有效问题'] == '是')
    ]['开发处理-开发所属部门'].value_counts()

    # 3. 统计每个组的缺陷问题量（按解决时间）
    defect_by_group = df[
        (solve_time_mask) & (df['缺陷问题'] == '是') 
    ]['开发处理-开发所属部门'].value_counts()

    # 4. 统计每个组的开发超期量（按解决时间）
    overdue_by_group = df[
        (solve_time_mask) & (df['开发是否超期'] == '是')
    ].apply(lambda x: x['超期责任人所属部门'] if pd.notna(x['超期责任人所属部门']) else x['开发处理-开发所属部门'], axis=1).value_counts()

    # 5. 统计每个组的有效问题处理时长（按解决时间）
    valid_issues = df[(solve_time_mask) & (df['有效问题'] == '是')]
    total_time_by_group = valid_issues.groupby('开发处理-开发所属部门')['开发处理-投入工作量'].sum()
    # 计算平均处理时长
    avg_time_by_group = valid_issues.groupby('开发处理-开发所属部门')['开发处理-投入工作量'].mean()

    # 6. 统计每个组的一次交付量（按解决时间）
    one_delivery_by_group = df[
        (solve_time_mask) & (df['一次交付'] == '是')
    ]['开发处理-开发所属部门'].value_counts()

    # 7. 统计每个组的严重问题量（按解决时间）
    serious_by_group = df[
        (solve_time_mask) & (df['严重问题'] == '是')
    ]['开发处理-开发所属部门'].value_counts()

    # 8. 统计每个组的处理完成量（按解决时间）
    completed_by_group = df[
        (solve_time_mask) & (df['开发处理-诊断结论-二级'].notna())
    ]['开发处理-开发所属部门'].value_counts()

    # 将结果存储在字典中
    result = {
        'total_count': total_by_group,
        'completed_count': completed_by_group,
        'valid_count': valid_by_group,
        'defect_count': defect_by_group,
        'overdue_count': overdue_by_group,
        'total_time': total_time_by_group,
        'avg_time': avg_time_by_group,
        'one_delivery_count': one_delivery_by_group,
        'serious_count': serious_by_group
    }

    return create_summary_dataframe(result, "部门")

def create_summary_dataframe(result, dimension_name):
    """创建汇总DataFrame"""
    # 创建一个DataFrame来显示所有统计结果
    summary_df = pd.DataFrame({
        '总问题量': result['total_count'],
        '处理完成量': result['completed_count'],
        '有效问题量': result['valid_count'],
        '缺陷问题量': result['defect_count'],
        '开发超期量': result['overdue_count'],
        '有效问题总处理时长(小时)': result['total_time'],
        '有效问题平均处理时长(小时)': result['avg_time'].round(2),
        '一次交付量': result['one_delivery_count'],
        '严重问题量': result['serious_count']
    }).fillna(0)

    # 计算百分比
    summary_df['有效率%'] = (summary_df['有效问题量'] / summary_df['总问题量'] * 100).round(2).astype(str)
    summary_df['缺陷率%'] = (summary_df['缺陷问题量'] / summary_df['总问题量'] * 100).round(2).astype(str)
    summary_df['处理及时率%'] = (100 - (summary_df['开发超期量'] / summary_df['处理完成量'] * 100)).round(2).astype(str)
    summary_df['一次交付率%'] = (summary_df['一次交付量'] / summary_df['处理完成量'] * 100).round(2).astype(str)
    summary_df['严重问题率%'] = (summary_df['严重问题量'] / summary_df['总问题量'] * 100).round(2).astype(str)

    # 按总问题量降序排列
    summary_df = summary_df.sort_values(by='总问题量', ascending=False)

    # 添加汇总行
    total_row = pd.DataFrame({
        '总问题量': [summary_df['总问题量'].sum()],
        '处理完成量': [summary_df['处理完成量'].sum()],
        '有效问题量': [summary_df['有效问题量'].sum()],
        '缺陷问题量': [summary_df['缺陷问题量'].sum()],
        '开发超期量': [summary_df['开发超期量'].sum()],
        '有效问题总处理时长(小时)': [summary_df['有效问题总处理时长(小时)'].sum()],
        '有效问题平均处理时长(小时)': [(summary_df['有效问题总处理时长(小时)'].sum() / summary_df['有效问题量'].sum()).round(2)],
        '一次交付量': [summary_df['一次交付量'].sum()],
        '严重问题量': [summary_df['严重问题量'].sum()]
    }, index=['总计'])

    # 计算汇总行的百分比
    total_row['有效率%'] = (total_row['有效问题量'] / total_row['总问题量'] * 100).round(2).astype(str)
    total_row['缺陷率%'] = (total_row['缺陷问题量'] / total_row['总问题量'] * 100).round(2).astype(str)
    total_row['处理及时率%'] = (100 - (total_row['开发超期量'] / total_row['处理完成量'] * 100)).round(2).astype(str)
    total_row['一次交付率%'] = (total_row['一次交付量'] / total_row['处理完成量'] * 100).round(2).astype(str)
    total_row['严重问题率%'] = (total_row['严重问题量'] / total_row['总问题量'] * 100).round(2).astype(str)

    # 将汇总行添加到主DataFrame
    summary_df = pd.concat([summary_df, total_row])

    # 打印表格形式的统计结果
    print(f"\n{dimension_name}问题统计报告:")
    print("-" * 50)
    print(summary_df)
    
    return summary_df

def main():
    """主函数"""
    # 解析命令行参数
    args = parse_args()
    
    # 设置Excel文件路径和时间区间
    EXCEL_FILE = args.excel_file
    START_DATE = args.start_date
    END_DATE = args.end_date
    
    print("Excel数据分析工具 (支持时间区间过滤)")
    print("=" * 60)
    print(f"正在分析文件: {EXCEL_FILE}")
    print(f"时间区间: {START_DATE} 至 {END_DATE}")
    
    try:
        # 读取Excel文件
        df = pd.read_excel(EXCEL_FILE)
        print(f"成功读取数据，共 {len(df)} 条记录")
        
        # 将时间列转换为datetime类型
        df['TB_上报时间'] = pd.to_datetime(df['TB_上报时间'])
        df['开发处理_开发解决完成时间'] = pd.to_datetime(df['开发处理_开发解决完成时间'])

        # 创建上报时间过滤条件
        report_time_mask = (df['TB_上报时间'] >= START_DATE) & (df['TB_上报时间'] <= END_DATE)
        
        # 创建解决时间过滤条件
        solve_time_mask = (df['开发处理_开发解决完成时间'] >= START_DATE) & (df['开发处理_开发解决完成时间'] <= END_DATE)
        
        print(f"上报时间在区间内的记录: {report_time_mask.sum()} 条")
        print(f"解决时间在区间内的记录: {solve_time_mask.sum()} 条")
        
        # 设置表格显示选项
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        
        # 执行三个维度的分析
        customer_df = analyze_by_customer(df, report_time_mask, solve_time_mask)
        developer_df = analyze_by_developer(df, report_time_mask, solve_time_mask)
        department_df = analyze_by_department(df, report_time_mask, solve_time_mask)
        
        # 将分析结果写入Excel文件
        print("\n正在将分析结果写入Excel文件...")
        
        # 加载工作簿
        wb = load_workbook(EXCEL_FILE)
        
        # 删除已存在的分析结果sheet
        sheets_to_remove = ['周分析结果', '周按人', '周分析结果(按部门)']
        for sheet_name in sheets_to_remove:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
        wb.save(EXCEL_FILE)
        
        # 写入新的分析结果
        with pd.ExcelWriter(EXCEL_FILE, mode='a', engine='openpyxl') as writer:
            # 写入客户维度分析结果
            customer_df.to_excel(writer, sheet_name='周分析结果')
            format_excel_worksheet(writer.sheets['周分析结果'], customer_df)
            
            # 写入开发人员维度分析结果
            developer_df.to_excel(writer, sheet_name='周按人')
            format_excel_worksheet(writer.sheets['周按人'], developer_df)
            
            # 写入部门维度分析结果
            department_df.to_excel(writer, sheet_name='周分析结果(按部门)')
            format_excel_worksheet(writer.sheets['周分析结果(按部门)'], department_df)
        
        print("分析完成！结果已保存到Excel文件中。")
        print("生成的工作表:")
        print("- 周分析结果 (按客户)")
        print("- 周按人 (按开发人员)")
        print("- 周分析结果(按部门) (按部门)")
        print(f"\n时间区间: {START_DATE} 至 {END_DATE}")
        
    except FileNotFoundError:
        print(f"错误：找不到文件 {EXCEL_FILE}")
        print("请检查文件路径是否正确")
    except Exception as e:
        print(f"分析过程中出现错误: {str(e)}")

if __name__ == "__main__":
    main() 