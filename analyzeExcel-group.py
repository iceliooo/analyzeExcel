import pandas as pd

from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 设置Excel文件路径
# EXCEL_FILE = 'file/客户问题上报 - 中文.xlsx'
# EXCEL_FILE = 'file/0330V8客户问题上报-分析.xlsx'
#EXCEL_FILE = 'file/0330V8-3.12客户问题上报-分析.xlsx'
# EXCEL_FILE = 'file/0331V8客户问题上报-分析.xlsx'
# EXCEL_FILE = 'file/0331/0331V8-3.12客户问题上报-分析.xlsx'
# EXCEL_FILE = 'file/0331/03月V8-3.12客户问题上报-分析.xlsx'
# EXCEL_FILE = 'file/0331/03月V8客户问题上报-分析.xlsx'

# EXCEL_FILE = 'file/0404V8客户问题上报 - 分析.xlsx'
# EXCEL_FILE = 'file/0411/0411V8客户问题上报 - 分析.xlsx'
# EXCEL_FILE = 'file/0418/0418V8客户问题上报 - 分析.xlsx'
# EXCEL_FILE = 'file/0425/0425V8客户问题上报 - 分析.xlsx'
# EXCEL_FILE = 'file/0430/0430V8客户问题上报 - 分析.xlsx'
# EXCEL_FILE = 'file/0509/0509V8客户问题上报 - 分析.xlsx'
# EXCEL_FILE = 'file/0516/0516V8客户问题上报 - 分析.xlsx'
# EXCEL_FILE = 'file/0523/0523V8客户问题上报 - 分析.xlsx'
# EXCEL_FILE = 'file/0613/0613V8客户问题上报 - 分析.xlsx'
# EXCEL_FILE = 'file/0627/0627V8客户问题上报 - 分析.xlsx'
EXCEL_FILE = 'file/0704/0704V8客户问题上报 - 分析.xlsx'

# EXCEL_FILE = 'file/V3.15/V3.15V8客户问题上报-分析V2.xlsx'
# EXCEL_FILE = 'file/V3.15/V3.15V8客户问题上报 (0426).xlsx'
# EXCEL_FILE = 'file/V3.18/V3.18V8客户问题上报  (0622).xlsx'

# EXCEL_FILE = 'file/V3.15/V3.15V8客户问题上报 (0430).xlsx'
# EXCEL_FILE = 'file/V3.15/V3.15V8客户问题上报 (0511).xlsx'
# EXCEL_FILE = 'file/V3.15/V3.15V8客户问题上报 (0518).xlsx'
# EXCEL_FILE = 'file/V3.15/V3.15V8客户问题上报 (0525).xlsx'
# EXCEL_FILE = 'file/V3.15/V3.15V8客户问题上报 (0531).xlsx'
# EXCEL_FILE = 'file/V3.18/V3.18V8客户问题上报 (0629).xlsx'
# EXCEL_FILE = 'file/V5.0/V5.0V8客户问题上报 (0629).xlsx'

# 读取Excel文件
df = pd.read_excel(EXCEL_FILE)

# 将开发处理-开发所属部门的空值替换为"未知"
df['开发处理-开发所属部门'] = df['开发处理-开发所属部门'].fillna('未知')

# 按研发组分组统计
# 1. 统计每个组的总问题量
total_by_group = df['开发处理-开发所属部门'].value_counts()

# 2. 统计每个组的有效问题量
valid_by_group = df[
    df['有效问题'] == '是'
]['开发处理-开发所属部门'].value_counts()

# 3. 统计每个组的缺陷问题量
defect_by_group = df[
    (df['缺陷问题'] == '是') 
]['开发处理-开发所属部门'].value_counts()

# 4. 统计每个组的开发超期量
overdue_by_group = df[
    df['开发是否超期'] == '是'
].apply(lambda x: x['超期责任人所属部门'] if pd.notna(x['超期责任人所属部门']) else x['开发处理-开发所属部门'], axis=1).value_counts()

# 5. 统计每个组的有效问题处理时长
valid_issues = df[df['有效问题'] == '是']
total_time_by_group = valid_issues.groupby('开发处理-开发所属部门')['开发处理-投入工作量'].sum()
# 计算平均处理时长
avg_time_by_group = valid_issues.groupby('开发处理-开发所属部门')['开发处理-投入工作量'].mean()

# 6. 统计每个组的一次交付量
one_delivery_by_group = df[
    df['一次交付'] == '是'
]['开发处理-开发所属部门'].value_counts()

# 7. 统计每个组的严重问题量
serious_by_group = df[
    df['严重问题'] == '是'
]['开发处理-开发所属部门'].value_counts()

# 8. 统计每个组的处理完成量
completed_by_group = df[
    df['开发处理-诊断结论-二级'].notna()
]['开发处理-开发所属部门'].value_counts()

# 将结果存储在字典中
result = {
    'total_count': total_by_group,
    'completed_count': completed_by_group,
    'valid_count': valid_by_group,
    'defect_count': defect_by_group,
    'overdue_count': overdue_by_group,
    'total_time': total_time_by_group,
    'avg_time': avg_time_by_group,
    'one_delivery_count': one_delivery_by_group,
    'serious_count': serious_by_group
}

# 创建一个DataFrame来显示所有统计结果
summary_df = pd.DataFrame({
    '总问题量': result['total_count'],
    '处理完成量': result['completed_count'],
    '有效问题量': result['valid_count'],
    '缺陷问题量': result['defect_count'],
    '开发超期量': result['overdue_count'],
    '有效问题总处理时长(小时)': result['total_time'],
    '有效问题平均处理时长(小时)': result['avg_time'].round(2),
    '一次交付量': result['one_delivery_count'],
    '严重问题量': result['serious_count']
}).fillna(0)

# 计算百分比
summary_df['有效率%'] = (summary_df['有效问题量'] / summary_df['总问题量'] * 100).round(2).astype(str)
summary_df['缺陷率%'] = (summary_df['缺陷问题量'] / summary_df['总问题量'] * 100).round(2).astype(str)
summary_df['处理及时率%'] = (100 - (summary_df['开发超期量'] / summary_df['处理完成量'] * 100)).round(2).astype(str)
summary_df['一次交付率%'] = (summary_df['一次交付量'] / summary_df['处理完成量'] * 100).round(2).astype(str)
summary_df['严重问题率%'] = (summary_df['严重问题量'] / summary_df['总问题量'] * 100).round(2).astype(str)

# 按总问题量降序排列
summary_df = summary_df.sort_values(by='总问题量', ascending=False)

# 添加汇总行
total_row = pd.DataFrame({
    '总问题量': [summary_df['总问题量'].sum()],
    '处理完成量': [summary_df['处理完成量'].sum()],
    '有效问题量': [summary_df['有效问题量'].sum()],
    '缺陷问题量': [summary_df['缺陷问题量'].sum()],
    '开发超期量': [summary_df['开发超期量'].sum()],
    '有效问题总处理时长(小时)': [summary_df['有效问题总处理时长(小时)'].sum()],
    '有效问题平均处理时长(小时)': [(summary_df['有效问题总处理时长(小时)'].sum() / summary_df['有效问题量'].sum()).round(2)],
    '一次交付量': [summary_df['一次交付量'].sum()],
    '严重问题量': [summary_df['严重问题量'].sum()]
}, index=['总计'])

# 计算汇总行的百分比
total_row['有效率%'] = (total_row['有效问题量'] / total_row['总问题量'] * 100).round(2).astype(str)
total_row['缺陷率%'] = (total_row['缺陷问题量'] / total_row['总问题量'] * 100).round(2).astype(str)
total_row['处理及时率%'] = (100 - (total_row['开发超期量'] / total_row['处理完成量'] * 100)).round(2).astype(str)
total_row['一次交付率%'] = (total_row['一次交付量'] / total_row['处理完成量'] * 100).round(2).astype(str)
total_row['严重问题率%'] = (total_row['严重问题量'] / total_row['总问题量'] * 100).round(2).astype(str)

# 将汇总行添加到主DataFrame
summary_df = pd.concat([summary_df, total_row])

# 设置表格显示选项
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)

# 打印表格形式的统计结果
print("\n研发组问题统计报告:")
print("-" * 50)
print(summary_df)

# 将分析结果写入Excel文件
from openpyxl import load_workbook
wb = load_workbook(EXCEL_FILE)
if '分析结果(按部门)' in wb.sheetnames:
    # 如果sheet已存在,删除旧sheet
    wb.remove(wb['分析结果(按部门)'])
wb.save(EXCEL_FILE)
wb.close()

with pd.ExcelWriter(EXCEL_FILE, mode='a', engine='openpyxl') as writer:
    summary_df.to_excel(writer, sheet_name='分析结果(按部门)') 

    # 获取工作表
    worksheet = writer.sheets['分析结果(按部门)']
    
    # 定义填充颜色
    blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    orange_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
    
    # 定义换行对齐方式
    wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    # 需要换行显示的列
    wrap_columns = ['有效问题总处理时长(小时)', '有效问题平均处理时长(小时)']
    wrap_col_indexes = []
    
    # 获取需要换行的列索引
    for i, col_name in enumerate(summary_df.columns):
        if col_name in wrap_columns:
            wrap_col_indexes.append(i + 2)  # +2 因为Excel索引从1开始且有索引列
    
    # 为所有列标题设置蓝色背景和居中对齐
    for col in range(1, len(summary_df.columns) + 2):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置适应内容的列宽
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].auto_size = True
    
    # 为需要换行的列设置格式
    for col_idx in wrap_col_indexes:
        # 为标题设置换行对齐
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.alignment = wrap_alignment
        
        # 为所有数据行设置换行对齐
        for row in range(2, len(summary_df) + 2):  # +2 因为Excel有标题行和索引从1开始
            cell = worksheet.cell(row=row, column=col_idx)
            cell.alignment = wrap_alignment
    
    # 查找特定列的索引
    special_columns = ['有效问题平均处理时长(小时)', '处理及时率%', '一次交付率%', '严重问题率%']
    special_col_indexes = []
    
    # 获取列索引 (注意Excel中的列从1开始，且第1列是索引)
    for i, col_name in enumerate(summary_df.columns):
        if col_name in special_columns:
            special_col_indexes.append(i + 2)  # +2 因为Excel索引从1开始且有索引列
    
    # 为特定列设置橙色背景
    for col_idx in special_col_indexes:
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.fill = orange_fill
        
        # 为总计行的特定单元格设置橙色
        total_row_cell = worksheet.cell(row=len(summary_df) + 1, column=col_idx)
        total_row_cell.fill = orange_fill
    
    # 自动调整列宽以适应内容
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width
